import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill

df = pd.read_excel('lab5.xlsx')
df = pd.concat([df.iloc[2:14, 0:5], df.iloc[2:14, 6:25]], axis=1, ignore_index=True)
df.to_excel('1.xlsx', index=False, header=False)
wb = load_workbook('1.xlsx')
ws = wb.active
thin_border = Border(
    left=Side(style='thick', color='0000FF'),
    right=Side(style='thick', color='0000FF'),
    top=Side(style='thick', color='0000FF'),
    bottom=Side(style='thick', color='0000FF')
)
left_top_right = Border(
    left=Side(style='thick', color='0000FF'),
    right=Side(style='thick', color='0000FF'),
    top=Side(style='thick', color='0000FF'),
)
left_top = Border(
    left=Side(style='thick', color='0000FF'),
    top=Side(style='thick', color='0000FF'),
)
left_bottom = Border(
    left=Side(style='thick', color='0000FF'),
    bottom=Side(style='thick', color='0000FF'),
)
right_top = Border(
    right=Side(style='thick', color='0000FF'),
    top=Side(style='thick', color='0000FF'),
)
right_bottom = Border(
    right=Side(style='thick', color='0000FF'),
    bottom=Side(style='thick', color='0000FF'),
)
left_right = Border(
    left=Side(style='thick', color='0000FF'),
    right=Side(style='thick', color='0000FF'),
)
left_bottom_right = Border(
    left=Side(style='thick', color='0000FF'),
    right=Side(style='thick', color='0000FF'),
    bottom=Side(style='thick', color='0000FF')
)
top = Border(
    top=Side(style='thick', color='0000FF'),
)
bottom = Border(
    bottom=Side(style='thick', color='0000FF'),
)
left = Border(
    left=Side(style='thick', color='0000FF'),
)
right = Border(
    right=Side(style='thick', color='0000FF'),
)
for col in range(1, 25):
    for row in range(1, 13):
        if 3 >= col >= 1 == row:
            border = left_top_right
        elif 1 <= col <= 3 and row == 12:
            border = left_bottom_right
        elif 1 <= col <= 3:
            border = left_right
        elif col == 4:
            if row == 1:
                border = top
            elif row == 12:
                border = bottom
            else:
                border = left
        elif col == 5:
            if row == 1:
                border = top
            elif row == 12:
                border = bottom
            else:
                border = right
        elif col == 6:
            if row == 1:
                border = left_top
            elif row == 12:
                border = left_bottom
            else:
                border = left
        elif col == 24:
            if row == 1:
                border = right_top
            elif row == 12:
                border = right_bottom
            else:
                border = right
        elif row == 1:
            border = top
        elif row == 12:
            border = bottom
        else:
            continue
        ws.cell(row=row, column=col).border = border
for col in ws.columns:
    mx = 0
    value = col[0].column_letter
    for cell in col:
        if cell.value:
            cell_length = len(str(cell.value))
            if cell_length > mx:
                mx = cell_length
    ws.column_dimensions[value].width = min(mx + 2, 40)
wb.save('1.xlsx')